import pandas as pd
import pdfplumber
import re
from openpyxl import Workbook
from openpyxl.styles import Font

def extract_numbers_from_text(text):
    """Extract all numbers from text using regular expressions"""
    # Pattern matches:
    # - Currency ($1,000.00)
    # - Decimals (3.14)
    # - Integers (42)
    # - Percentages (15%)
    # - Negative numbers (-5.2)
    number_pattern = r"""
        (?:^|\s)          # Start of string or whitespace
        ([-+]?            # Optional sign
        \$?\d{1,3}        # Optional dollar sign and 1-3 digits
        (?:,\d{3})*       # Optional thousands separators
        (?:\.\d+)?        # Optional decimal portion
        \%?)              # Optional percent sign
        (?=\s|$)          # Lookahead for whitespace or end
    """
    return re.findall(number_pattern, text, re.VERBOSE)

def pdf_to_excel_with_numbers(pdf_path, excel_path):
    """Extract text, tables, and numbers from PDF to Excel with separate numbers section"""
    try:
        # Initialize workbook with styled fonts
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        
        # Formatting styles
        header_font = Font(bold=True, color="FF0000")
        number_font = Font(color="0000FF")  # Blue for numbers
        
        current_row = 1
        all_numbers = []  # Store all extracted numbers
        
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                # ========== TEXT EXTRACTION ==========
                text = page.extract_text()
                if text:
                    # Write text header
                    ws.cell(row=current_row, column=1, 
                          value=f"=== Page {page_num} Text ===").font = header_font
                    current_row += 1
                    
                    # Process text lines
                    for line in text.split('\n'):
                        if line.strip():
                            # Write text line
                            ws.cell(row=current_row, column=1, value=line.strip())
                            
                            # Extract numbers from this line
                            line_numbers = extract_numbers_from_text(line)
                            if line_numbers:
                                all_numbers.extend(line_numbers)
                            
                            current_row += 1
                    
                    current_row += 1  # Add spacing
                
                # ========== TABLE EXTRACTION ==========
                tables = page.extract_tables()
                for table_num, table in enumerate(tables, 1):
                    if table and len(table) > 1:
                        try:
                            df = pd.DataFrame(table[1:], columns=table[0])
                            
                            # Write table header
                            ws.cell(row=current_row, column=1, 
                                  value=f"=== Page {page_num} Table {table_num} ===").font = header_font
                            current_row += 1
                            
                            # Write headers
                            for col_num, header in enumerate(df.columns, 1):
                                ws.cell(row=current_row, column=col_num, 
                                      value=str(header).strip())
                            
                            current_row += 1
                            
                            # Write table data and extract numbers
                            for _, row in df.iterrows():
                                for col_num, value in enumerate(row, 1):
                                    cell_value = str(value).strip() if pd.notna(value) else ""
                                    ws.cell(row=current_row, column=col_num, value=cell_value)
                                    
                                    # Extract numbers from table cells
                                    if cell_value:
                                        cell_numbers = extract_numbers_from_text(cell_value)
                                        if cell_numbers:
                                            all_numbers.extend(cell_numbers)
                                
                                current_row += 1
                            
                            current_row += 1  # Add spacing
                            
                        except Exception as e:
                            print(f"Error processing table {table_num} on page {page_num}: {str(e)}")
        
        # ========== NUMBERS SECTION ==========
        if all_numbers:
            # Write numbers header
            ws.cell(row=current_row, column=1, 
                  value="=== Extracted Numbers ===").font = header_font
            current_row += 1
            
            # Write numbers in 5 columns for better readability
            num_cols = 5
            for i, num in enumerate(all_numbers):
                row = current_row + (i // num_cols)
                col = 1 + (i % num_cols)
                ws.cell(row=row, column=col, value=num).font = number_font
            
            current_row += len(all_numbers) // num_cols + 2
        
        # Save workbook
        wb.save(excel_path)
        print(f"Successfully extracted {len(all_numbers)} numbers to {excel_path}")
        return True
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return False

# Example usage
pdf_path = "sample.pdf"  # Replace with your PDF
excel_path = "output_with_numbers.xlsx"
pdf_to_excel_with_numbers(pdf_path, excel_path)
